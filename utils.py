import io
import os
import shutil
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image, HRFlowable, KeepTogether
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

class NumberedCanvas:
    """Two-pass canvas to dynamically compute and print 'Page X sur Y' and header/footer."""
    def __init__(self, *args, **kwargs):
        self._saved_page_states = []

    def __call__(self, *args, **kwargs):
        from reportlab.pdfgen import canvas
        class PageNumCanvas(canvas.Canvas):
            def __init__(self, *c_args, **c_kwargs):
                super().__init__(*c_args, **c_kwargs)
                self.pages = []

            def showPage(self):
                self.pages.append(dict(self.__dict__))
                self._startPage()

            def save(self):
                num_pages = len(self.pages)
                for page in self.pages:
                    self.__dict__.update(page)
                    self.draw_page_decorations(num_pages)
                    super().showPage()
                super().save()

            def draw_page_decorations(self, page_count):
                self.saveState()
                self.setFont("Helvetica", 9)
                self.setFillColor(colors.HexColor("#6c757d"))
                
                # Header
                self.drawString(54, 842 - 36, "Mon Suivi Financier - Rapport Financier")
                self.drawRightString(595 - 54, 842 - 36, datetime.now().strftime("Généré le %d/%m/%Y à %H:%M"))
                self.setStrokeColor(colors.HexColor("#e0e0e0"))
                self.setLineWidth(0.5)
                self.line(54, 842 - 42, 595 - 54, 842 - 42)

                # Footer
                self.line(54, 45, 595 - 54, 45)
                self.drawString(54, 30, "Document confidentiel - Généré automatiquement")
                page_text = f"Page {self._pageNumber} sur {page_count}"
                self.drawRightString(595 - 54, 30, page_text)
                self.restoreState()

        return PageNumCanvas(*args, **kwargs)

def generate_excel_export(transactions, title="Export_Transactions", currency="DH", is_monthly_report=False):
    """Generate Excel binary buffer with bold headers, auto column width, logo header, and styled formatting."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"
    ws.views.sheetView[0].showGridLines = True

    # Insert Logo Image if available
    logo_path = os.path.join(os.path.dirname(__file__), 'static', 'img', 'logo.png')
    if os.path.exists(logo_path):
        try:
            from openpyxl.drawing.image import Image as OpenPyxlImage
            xl_img = OpenPyxlImage(logo_path)
            xl_img.width = 42
            xl_img.height = 42
            ws.add_image(xl_img, "A1")
        except Exception:
            pass

    ws.row_dimensions[1].height = 35
    ws["B1"] = "Personal Expense Tracker"
    ws["B1"].font = Font(name="Calibri", size=14, bold=True, color="0D6EFD")
    ws["B1"].alignment = Alignment(vertical="center")

    ws.append([]) # Row 2 blank

    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0D6EFD", end_color="0D6EFD", fill_type="solid")
    border_side = Side(style="thin", color="E0E0E0")
    cell_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    
    number_format = f'#,##0.00 "{currency}"'

    if is_monthly_report:
        headers = ["Date", "Description", f"Revenu ({currency})", f"Dépense ({currency})", f"Net ({currency})"]
    else:
        headers = ["Date", "Description", f"Revenu ({currency})", f"Dépense ({currency})"]

    ws.append(headers) # Row 3 Headers

    header_row_idx = 3
    for cell in ws[header_row_idx]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    tot_rev = 0.0
    tot_dep = 0.0

    for item in transactions:
        d_str = item.date.strftime("%d/%m/%Y") if hasattr(item.date, "strftime") else str(item.date)
        rev = float(item.revenu or 0.0)
        dep = float(item.depense or 0.0)
        tot_rev += rev
        tot_dep += dep

        if is_monthly_report:
            net = rev - dep
            row_data = [d_str, item.description or "-", rev, dep, net]
        else:
            row_data = [d_str, item.description or "-", rev, dep]

        ws.append(row_data)

    # Style data rows
    for row in ws.iter_rows(min_row=header_row_idx + 1, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for idx, cell in enumerate(row):
            cell.border = cell_border
            if idx in (2, 3, 4) if is_monthly_report else (2, 3):
                cell.number_format = number_format
                cell.alignment = Alignment(horizontal="right")
            elif idx == 0:
                cell.alignment = Alignment(horizontal="center")
            else:
                cell.alignment = Alignment(horizontal="left")

    # Add Summary Row
    ws.append([]) # Empty separator
    if is_monthly_report:
        summary_row = ["TOTAL", "Total général", tot_rev, tot_dep, tot_rev - tot_dep]
    else:
        summary_row = ["TOTAL", f"{len(transactions)} transaction(s)", tot_rev, tot_dep]
    
    ws.append(summary_row)
    sum_row_num = ws.max_row

    sum_font = Font(name="Calibri", size=11, bold=True, color="000000")
    sum_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
    sum_border = Border(top=Side(style="medium", color="000000"), bottom=Side(style="double", color="000000"))

    for cell in ws[sum_row_num]:
        if cell.value is not None:
            cell.font = sum_font
            cell.fill = sum_fill
            cell.border = sum_border
            if isinstance(cell.value, (int, float)):
                cell.number_format = number_format

    # Auto column width
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def generate_pdf_report(transactions, month_str, year_str, app_name="Mon Suivi Financier", currency="DH", logo_path=None, include_summary=False):
    """Generate a clean, professional PDF report using ReportLab with static/img/logo.png header.
       include_summary (bool): Controls whether the summary cards block is rendered in PDF. Default is False.
    """
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=54,
        rightMargin=54,
        topMargin=54,
        bottomMargin=54
    )

    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=18,
        leading=22,
        textColor=colors.HexColor('#0D6EFD')
    )
    
    subtitle_style = ParagraphStyle(
        'DocSubTitle',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=10,
        leading=13,
        textColor=colors.HexColor('#6C757D')
    )

    table_header_style = ParagraphStyle(
        'TableHeader',
        fontName='Helvetica-Bold',
        fontSize=9,
        leading=11,
        textColor=colors.white,
        alignment=1 # Center
    )

    table_cell_style = ParagraphStyle(
        'TableCell',
        fontName='Helvetica',
        fontSize=9,
        leading=11,
        textColor=colors.HexColor('#212529')
    )

    table_cell_bold = ParagraphStyle(
        'TableCellBold',
        fontName='Helvetica-Bold',
        fontSize=9,
        leading=11,
        textColor=colors.HexColor('#212529')
    )

    summary_val_style = ParagraphStyle(
        'DocSummaryVal',
        fontName='Helvetica-Bold',
        fontSize=12,
        leading=15,
        alignment=1 # Center
    )

    story = []

    # Header block with logo.png
    resolved_logo_path = os.path.join(os.path.dirname(__file__), 'static', 'img', 'logo.png')
    if os.path.exists(resolved_logo_path):
        logo_cell = Image(resolved_logo_path, width=1.5*cm, height=1.5*cm)
        header_data = [
            [
                logo_cell,
                Paragraph(f"<b>{app_name}</b>", title_style),
                Paragraph(f"<b>Rapport Financier Mensuel</b><br/>Période: {month_str} {year_str}", subtitle_style)
            ]
        ]
        header_table = Table(header_data, colWidths=[50, 230, 207])
    else:
        header_data = [
            [
                Paragraph(f"<b>{app_name}</b>", title_style),
                Paragraph(f"<b>Rapport Financier Mensuel</b><br/>Période: {month_str} {year_str}", subtitle_style)
            ]
        ]
        header_table = Table(header_data, colWidths=[280, 207])

    header_table.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('ALIGN', (-1,0), (-1,0), 'RIGHT'),
    ]))
    story.append(header_table)
    story.append(Spacer(1, 15))
    story.append(HRFlowable(width="100%", thickness=1.5, color=colors.HexColor('#0D6EFD'), spaceAfter=15))

    tot_rev = sum(float(t.revenu or 0.0) for t in transactions)
    tot_dep = sum(float(t.depense or 0.0) for t in transactions)
    net_mois = tot_rev - tot_dep

    # Optional Summary Cards Box in PDF
    if include_summary:
        summary_data = [
            [
                Paragraph("<b>Total Revenus</b>", table_cell_bold),
                Paragraph("<b>Total Dépenses</b>", table_cell_bold),
                Paragraph("<b>Net du Mois</b>", table_cell_bold),
                Paragraph("<b>Transactions</b>", table_cell_bold)
            ],
            [
                Paragraph(f"<font color='#198754'><b>+{tot_rev:,.2f} {currency}</b></font>", summary_val_style),
                Paragraph(f"<font color='#DC3545'><b>-{tot_dep:,.2f} {currency}</b></font>", summary_val_style),
                Paragraph(f"<font color='{'#198754' if net_mois >= 0 else '#DC3545'}'><b>{net_mois:,.2f} {currency}</b></font>", summary_val_style),
                Paragraph(f"<b>{len(transactions)}</b>", summary_val_style)
            ]
        ]

        summary_table = Table(summary_data, colWidths=[120, 120, 140, 107])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#F8F9FA')),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('BOX', (0,0), (-1,-1), 1, colors.HexColor('#DEE2E6')),
            ('INNERGRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E9ECEF')),
            ('TOPPADDING', (0,0), (-1,-1), 8),
            ('BOTTOMPADDING', (0,0), (-1,-1), 8),
        ]))
        story.append(summary_table)
        story.append(Spacer(1, 20))

    # Transactions Table
    table_rows = [
        [
            Paragraph("Date", table_header_style),
            Paragraph("Description", table_header_style),
            Paragraph(f"Revenu ({currency})", table_header_style),
            Paragraph(f"Dépense ({currency})", table_header_style),
            Paragraph(f"Net ({currency})", table_header_style)
        ]
    ]

    for t in transactions:
        d_str = t.date.strftime("%d/%m/%Y") if hasattr(t.date, "strftime") else str(t.date)
        rev = float(t.revenu or 0.0)
        dep = float(t.depense or 0.0)
        net = rev - dep

        rev_text = f"<font color='#198754'>+{rev:,.2f}</font>" if rev > 0 else "0.00"
        dep_text = f"<font color='#DC3545'>-{dep:,.2f}</font>" if dep > 0 else "0.00"
        net_color = '#198754' if net > 0 else ('#DC3545' if net < 0 else '#212529')
        net_text = f"<font color='{net_color}'><b>{net:+,.2f}</b></font>"

        table_rows.append([
            Paragraph(d_str, table_cell_style),
            Paragraph(t.description or "-", table_cell_style),
            Paragraph(rev_text, table_cell_style),
            Paragraph(dep_text, table_cell_style),
            Paragraph(net_text, table_cell_style)
        ])

    # Add Summary Row to Table
    table_rows.append([
        Paragraph("<b>TOTAL</b>", table_cell_bold),
        Paragraph(f"<b>{len(transactions)} transaction(s)</b>", table_cell_bold),
        Paragraph(f"<font color='#198754'><b>+{tot_rev:,.2f}</b></font>", table_cell_bold),
        Paragraph(f"<font color='#DC3545'><b>-{tot_dep:,.2f}</b></font>", table_cell_bold),
        Paragraph(f"<font color='{'#198754' if net_mois>=0 else '#DC3545'}'><b>{net_mois:+,.2f}</b></font>", table_cell_bold)
    ])

    pdf_table = Table(table_rows, colWidths=[70, 207, 70, 70, 70])
    
    t_style = [
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#0D6EFD')),
        ('ALIGN', (0,0), (0,-1), 'CENTER'),
        ('ALIGN', (2,0), (-1,-1), 'RIGHT'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('GRID', (0,0), (-1,-2), 0.5, colors.HexColor('#E0E0E0')),
        ('BACKGROUND', (0,-1), (-1,-1), colors.HexColor('#E9ECEF')),
        ('LINEABOVE', (0,-1), (-1,-1), 1.5, colors.HexColor('#0D6EFD')),
    ]

    # Alternating row colors
    for i in range(1, len(table_rows) - 1):
        if i % 2 == 0:
            t_style.append(('BACKGROUND', (0, i), (-1, i), colors.HexColor('#F8F9FA')))

    pdf_table.setStyle(TableStyle(t_style))
    story.append(pdf_table)

    # Build PDF
    canvas_factory = NumberedCanvas()
    doc.build(story, canvasmaker=canvas_factory)
    buffer.seek(0)
    return buffer
